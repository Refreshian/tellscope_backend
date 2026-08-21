from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from .catalog import Catalog
from .models import ReportBundle


def export_excel(bundle: ReportBundle, catalog: Catalog, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "objects"
    ws.append(
        [
            "id", "short", "kind", "messages", "main_role", "episodic", "background",
            "positive", "negative", "neutral", "initiated", "unique_media",
            "top2_share", "reach", "media_index", "from_vendor",
        ]
    )
    for st in bundle.object_stats:
        obj = catalog.by_id.get(st.object_id)
        ws.append(
            [
                st.object_id,
                obj.short if obj else st.object_id,
                obj.kind if obj else "",
                st.messages,
                st.main_role,
                st.episodic_role,
                st.background_role,
                st.positive,
                st.negative,
                st.neutral,
                st.initiated,
                st.unique_media,
                round(st.top2_share, 4),
                st.reach,
                st.media_index,
                st.from_vendor,
            ]
        )

    ws = wb.create_sheet("messages")
    ws.append(
        [
            "id", "published_at", "source", "contour", "title", "url",
            "objects", "speakers", "sentiment", "initiated", "classified_by", "file",
        ]
    )
    for msg in bundle.messages:
        ws.append(
            [
                msg.id,
                msg.published_at.isoformat(sep=" ") if msg.published_at else "",
                msg.source,
                msg.contour,
                msg.title,
                msg.url,
                ", ".join(catalog.label(i) for i in msg.object_ids),
                ", ".join(
                    catalog.person_by_id[i].short if i in catalog.person_by_id else i
                    for i in msg.speaker_ids
                ),
                msg.sentiment,
                msg.initiated,
                msg.classified_by,
                msg.file_name,
            ]
        )

    ws = wb.create_sheet("media")
    ws.append(["object", "outlet", "messages"])
    for st in bundle.object_stats:
        for outlet, n in sorted(st.media.items(), key=lambda x: -x[1]):
            ws.append([catalog.label(st.object_id), outlet, n])

    ws = wb.create_sheet("speakers")
    ws.append(["object", "speaker", "messages"])
    for st in bundle.object_stats:
        for sp, n in sorted(st.speakers.items(), key=lambda x: -x[1]):
            label = catalog.person_by_id[sp].short if sp in catalog.person_by_id else sp
            ws.append([catalog.label(st.object_id), label, n])

    ws = wb.create_sheet("daily")
    ws.append(["object", "date", "messages"])
    for st in bundle.object_stats:
        for day, n in sorted(st.daily.items()):
            ws.append([catalog.label(st.object_id), day, n])

    ws = wb.create_sheet("notes")
    ws.append(["note"])
    for note in bundle.notes:
        ws.append([note])
    for miss in bundle.missing_metrics:
        ws.append([f"нет метрики: {miss}"])

    wb.save(path)
